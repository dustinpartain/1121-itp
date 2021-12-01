# ITP Week 3 Day 1 Lecture

## Dustin Notes
## 2 stages of functions

## first: declaring (def)
## second: invocation (calling)

## different types of function

## void
## return (with or without parameters)

## modules = reuseable 'packages' of code that can be pulled into own code... i.e. Python package of "Math module"

#------------PIP-----------

#------------EXCEL SPREADSHEET-----------

# -Workbook
#     -can contain multiple sheets (worksheets)
# -xlsx file extension
# -Sheets  / Worksheets
# -Columns (letters) & Rows (numbers)
# -Cell
#     -intersection of row & Column

# READ
# - SHEET, CELLS, COLUMNS, ROWS

#------------OPENPYXL-----------
# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).

# -openpyxl is a downloaded module
# -Download & install openpyxl using pip
#   pip install openpyxl

import openpyxl

#Assign your .xlsx file to a variable
# filename = "path/to/your/excel"
openpyxl.load_workbook(filename) #returns a Workbook object.
my_workbook = openpyxl.load_workbook('example.xlsx')
#use the 'type' method on the new variable to verify what kind of data type you are working with
type(my_workbook) # Result-->  <class 'openpyxl.workbook.workbook.Workbook'>



#Each Excel file can have multiple sheets, which each contain many rows, columns and cells.  You must specify which sheet you would like to manipulate.
get_sheet_names() #and
get_sheet_by_name() #help get Worksheet objects.

my_workbook.get_sheet_names()  # Result-->  ['Sheet1', 'Sheet2', 'Sheet3']

my_sheet = my_workbook.get_sheet_by_name('Sheet 1')
#verify
type(my_sheet)  # Result-->  <class 'openpyxl.worksheet.worksheet.Worksheet'>

# Cell objects have a "value" member variable with the content of that cell.
# The square brackets in sheet['A1'] get Cell objects.
# The cell() method also returns a Cell object from a sheet.
my_sheet['A1'] #will return the CELL OBJECT (NOT the value you see in the cell) created by the intersection of the first column and the first row.  The reult of sheet['A1'] will actually be:
    # Result -->  <Cell Sheet1.A1>
#so we must ASSIGN IT TO A VARIABLE before we can work with it.
my_cell = sheet['A1']
my_cell.value() # Result --> the value of the A1 cell

#Don't forget to convert the values to a string if you are working with strings
str(my_cell.value)
str(my_sheet['A1'].value)

#You can also access cell values using the .cell() method and passing arguments for the column and row:
my_sheet.cell(row=1, column=2) # Result --> <Cell Sheet1.B1>


#Python can be used to loop through the cells on your spreadsheet and manipulate data automatically.
for i in range(1, 8):
    print(i, my_sheet.cell(row = i, column =2).value)
#The code may be difficult to understand at first, but notice that in 'row = i', the 'i' will increase by 1 until it reaches 8, thus printing out each value in column 2.  The first 'i', after the 'print(' is to show the correlation between 'i' and row being printed.


# READ DATA

# WRITE TO A DATA

# UPDATE A SPREADSHEET

# SET UP A DATA STRUCTURE WITH THE UPDATED INFORMATION

# FORMULAS

# OTHER FUNCTIONALITIES


######################################
#-----------GIT & GITHUB-------------
######################################

#     GIT IS NOT THE SAME AS GITHUB!!!!!!!!!

#A .git file is created in your folder when git is initialized (the command to do a simple git intializtion is "git init", but we're not using that today).

#GitHUB is an online service which hosts code IN TEXT FORM.  It does NOT run any programs.  It is used by individuals to save their code to the cloud, and teams use it to work on large applications where individuals are responsible for different parts.

#Your computer has a lot of files.  You don't want to upload ALL of them, so you use the "git add ." from within the folder you would like Git to track. (The "." indicates "all of the files in the folder")

#After you have told your computer which files to keep track of, you will need to commit your files in order to prepare them to be sent to github.  This is also the point that you can think of as "saving your game".  You will need to add a message describing the changes you've made, so the syntax will look like:
#git commit -m "I just changed some stuff"

#After your changes have been saved, you may now push it to github using the "git push" command 

#---------HOW TO USE OTHER CODE---------
#Go to the repository on github that you would like to copy.  In our case it will be:
# https://github.com/Vets-In-Tech/itp_week_2

# In the top right corner, select the "fork" button.
# Once the repository has been copied to your account, go to it and click on the green "Code" button.  A dropdown menu will appear.  Select the clipboard.

#Back in your terminal, open the directory you would like to clone the repository to.  Enter:
#git clone https://github.com/Vets-In-Tech/itp_week_2

#This should clone the repository.  Now you know how to fork, clone, add, commit, and push.  But what about updating your local files with the changes from github?  The command is:
#git pull

#You can also use the git tools in the lefthand toolbar (it's the icon with the three circles and two lines)


